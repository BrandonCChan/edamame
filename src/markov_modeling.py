#---------------------------------------------------------------------------------------------------
# Code for running a monte carlo markov using a model structure and parameters specified in an excel
# document. (See README and test_parameters.xlsx for addional explaination and example)
#
# Brandon Chan | July 2020
#---------------------------------------------------------------------------------------------------
# Import packages/libraries
#---------------------------------------------------------------------------------------------------
import pandas as pd # Dataframe structure and manipulation
import numpy as np # Scientific computing functionality (array and matrix operations and some stats things)
import math # For additional math functions
from time import perf_counter, strftime # For timing the runtime of the code
from openpyxl import load_workbook # Helps interface with excel document for read/writing
from model_helpers import * # Importing all the functions outlined in model_helpers.py

def run_model(filepath, save=False, model_name='model'):
    '''
    Function that reads a specified excel sheet and runs the model specified by it
    Input: filepath = path of excel doc that defines model
    Optional inputs: save = boolean (True/False) flag to automatically save the model outputs in the
                            model_outputs/ directory. Default is Flase
                     model_name = Used when save is equal to True. Specifies a string to be used to
                                  as a in-filename descriptor for the saved model outputs. Default is "model" 
    Output: returns 3 npy arrays of the population, cost, and utility over each model iteration, 
            state, and cycle.
    '''
    #---------------------------------------------------------------------------------------------------
    # File IO and parameter initialization
    #---------------------------------------------------------------------------------------------------
    input_file = pd.ExcelFile(filepath) # read in excel file
    check_excel_file(input_file) # run checks on file formatting and specification of model

    # Read in each relevant speadsheet in the file
    transitions_df = pd.read_excel(input_file, 'transitions')
    costs_df = pd.read_excel(input_file, 'costs')
    utilities_df = pd.read_excel(input_file, 'utilities')
    specification_df = pd.read_excel(input_file, 'specification', header=None, index_col=0)

    # Specification of variables regarding states in model
    unique_states = transitions_df['start_state'].unique().tolist()
    num_states = len(unique_states)

    # Initialize model parameters based on spreadsheet defined values
    max_iterations = int(specification_df.loc['max_iterations'].values[0])
    cycle_length = specification_df.loc['cycle_length'].values[0]
    time_horizon_days = specification_df.loc['time_horizon'].values[0] * 365
    num_cycles = int(time_horizon_days / cycle_length)
    name_start_state = specification_df.loc['name_start_state'].values[0]
    discount_rate = specification_df.loc['discount_rate'].values[0]

    print('file and parameters loaded...')

    #---------------------------------------------------------------------------------------------------
    # Generate matrix representation of model and identify/log state transitions that require resampling
    # or additional calcuations when simulating
    # 
    # Dimensions of matrix = [num_states x num_states]
    # Rows map to starting state, columns map to target state
    #---------------------------------------------------------------------------------------------------
    # Use dict to map from names to numeric index in array 
    state_mapping = {i : unique_states.index(i) for i in unique_states}
    #specify empty transition matrix
    transition_matrix = np.zeros((num_states,num_states)) 

    # Keep track of specific indicies in the transition matrix that need to be updated "in-simulation"
    # ie. time-dependent transitions and transitions that get resampled/recalculated every iteration
    # Intended to be stored as a list of dictionaries. Dictionaries contain key-value pairs that denote
    # type of transtion, index (i,j) of matrix corresponding to transtion, and appropriate parameters ie. a, b, shape, scale, etc.
    resample_indicies = []
    time_dependent_indicies = []
    residual_indicies = []

    # Iterate through specified transtions and initialize constant values of matrix. Transitions that vary 
    # (ie. time dependent or resampled) are assigned in model iteration step. 
    # Log indicies of transitions that need to be updated to quickly index the correct position in the transition matrix
    for t in transitions_df.itertuples():
        start_state_index = state_mapping[t[1]] # mapped row number of start state
        end_state_index = state_mapping[t[2]] # mapped column number of end state
        t_type = t[3] # type of transition 
        params = t[4:] # parameters
        
        if t_type == 'constant':
            transition_matrix[start_state_index, end_state_index] = set_transition('constant', transition=params[0])
        elif t_type in ['beta', 'gamma']:
            resample_indicies += [{'start_state':t[1],'end_state':t[2],'i':start_state_index, 'j':end_state_index, 'type':t_type, 'a':params[0], 'b':params[1]}]
        elif t_type == 'time_dependent_weibull' or t_type == 'time_dependent_gompertz':
            time_dependent_indicies += [{'start_state':t[1],'end_state':t[2],'i':start_state_index, 'j':end_state_index, 'type':t_type, 'const':params[0], 'ancillary':params[1]}]
        elif t_type == 'residual':
            residual_indicies += [{'start_state':t[1],'end_state':t[2],'i':start_state_index, 'j':end_state_index, 'type':t_type, 'params':params}]
        else:
            raise ValueError('Invalid transition type provided:',str(t_type),'Please double check excel file specification')

    #---------------------------------------------------------------------------------------------------
    # Run the simulation (for a single arm)
    # TODO: could wrap this into a numba function for increase in speed
    #---------------------------------------------------------------------------------------------------
    # Create result logging array/dataframe 
    # Shape: [iteration_number x states x timesteps] where timesteps = number of cycles
    results_log = np.zeros((max_iterations, num_states, num_cycles))

    print('beginning iterations...')
    iteration_times = np.zeros((max_iterations,1))
    for iteration in range(0,max_iterations):
        start_iteration_time = perf_counter() # begin timer for logging run time of model iteration

        # Initialize the starting population/proportion for each state at beginning of each model iteration
        # I.e. starting with a zero column-vector of dimension [number_of_states x 1]
        # The initial "proportions" are assigned in accordance to the corresponding row of the vector
        # This corresponds with the row/name assignments of the transition matrix.
        population = np.zeros((num_states,1)) 
        population[state_mapping[name_start_state]] = 1 
        
        # Initialize with population at time 0 (ie. first cycle everyone is in tx_1 or whereever presumablly)
        results_log[iteration, :, 0] = population.reshape(num_states)
        
        # Initialize as 1 at every iteration becasue population at 0 is always the same at the beginning of
        # any given iteration?
        cycle = 1

        # Resample transition probailities if needed (ie. from distributions) - Update matrix as appropriate
        for t in resample_indicies:
            transition_matrix[t['i'],t['j']] = set_transition(t['type'], a=t['a'], b=t['b'])  
        
        # For every timestep until max is reached
        while cycle < num_cycles:
            # Adjust time-dependent transition probabilities based on timestep if needed
            for t in time_dependent_indicies:
                transition_matrix[t['i'],t['j']] = set_transition(t['type'], const=t['const'], ancillary=t['ancillary'], cycle=cycle, cycle_length=cycle_length)
            
            # Calculate residual transition probailities if needed - Update matrix as appropriate
            for t in residual_indicies:
                transition_matrix[t['i'], t['j']] = 0 # AFTER RESAMPLING RESIDUAL IS RECALCULATED FROM ZERO (but only zero specific transition i to j)
                transition_matrix[t['i'], t['j']] = calculate_residual(transition_matrix, t['i'])
       
            # Normalize to a valid matrix and perform sum check
            transition_matrix = normalize_transitions(transition_matrix) 
            check_row_sums(transition_matrix)
            
            # Calculate the movement for the "population" in each state to another
            # population is already a column vector, dont need to transpose.
            population = transition_matrix.T @ population
            
            # Check: does population sum to 1? (assuming a round to 5 significant digits hold)
            check_model_population(population)
            
            # Update results_log after ever timestep
            results_log[iteration, :, cycle] = population.reshape(num_states) 
            
            cycle += 1 # move to next cycle

        iteration_times[iteration] = [perf_counter() - start_iteration_time] # log time required to run interation

    print('model done...')
    print('total time:',round(iteration_times.sum(),2),'seconds || mean time per iteration:', round(iteration_times.mean(),2),'seconds') 

    # Save model output if flagged to do so. Moved up to gaurd against cost/util failure 
    if save:
        name = model_name + '_' + strftime("%d-%m-%Y")
        np.save('../model_outputs/'+name+'_population.npy', results_log)
        print('model output saved: ../model_outputs/'+name+'... .npy')

        # Save state mappings in a separate worksheet
        book = load_workbook(filepath)
        if 'state_mappings' not in book.sheetnames:
        # TODO:// add more powerful matching and logic. ie. update instead of ignore if different
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                writer.book = book
                state_mapping_df = pd.DataFrame.from_dict(state_mapping, orient='index')
                state_mapping_df.to_excel(writer, sheet_name='state_mappings')

    #---------------------------------------------------------------------------------------------------
    # Costing and utility component
    #---------------------------------------------------------------------------------------------------
    # TODO: implement substate calcultions/identification/specification in excel sheet
    # Empty dict intitalized to work with roughed in condition in downstream code
    multiple_toxicity_states = {}

    # Initialize results arrays
    results_log_costs = np.zeros(results_log.shape)
    results_log_utilities = np.zeros(results_log.shape)

    # For sampling costs/utilities per interation
    # ie. generate a [num_states, num_iterations] shaped array 
    # TODO: move these or at least add checks for these in the check_file function
    iteration_costs = np.zeros((results_log.shape[0], results_log.shape[1]))
    copy_costs = {}
    for t in costs_df.itertuples():
        state_index = state_mapping[t[1]]
        c_type = t[2]
        if c_type == 'beta':
            for i in range(0,iteration_costs.shape[0]):
                iteration_costs[:,state_index] = get_beta(t[3], t[4])
        elif c_type == 'gamma':
            for i in range(0,iteration_costs.shape[0]):
                iteration_costs[:,state_index] = get_gamma(t[3], t[4])
        elif c_type == 'static':
            iteration_costs[:,state_index] = t[3]
        elif c_type == 'copy':
            copy_costs[state_index] = [state_mapping[t[3]]] # cost for index (noted as copy) points to index of target
            iteration_costs[state_index] = np.nan # initialize as nan
        else:
            raise ValueError('Error: Bad cost type specification', c_type)

    # Fill copied costs
    for c in copy_costs:
        iteration_costs[c] = iteration_costs[copy_costs[c]]
    if np.isnan(iteration_costs).any(): #TODO: Could make this output something more detailed. I.e. index of error cost
        raise ValueError('Error: NaN cost specified. Likely a copy type error. Please check costs sheet')
    
    iteration_utils = np.zeros((results_log.shape[0], results_log.shape[1]))
    copy_utils = {}
    for t in utilities_df.itertuples():
        state_index = state_mapping[t[1]]
        u_type = t[2]
        if u_type == 'beta':
            for i in range(0,iteration_utils.shape[0]):
                iteration_utils[i,state_index] = get_beta(t[3], t[4])
        elif u_type == 'gamma':
            for i in range(0,iteration_utils.shape[0]):
                iteration_utils[i,state_index] = get_gamma(t[3], t[4])
        elif u_type == 'static':
            iteration_utils[:,state_index] = t[3]
        elif u_type == 'copy':
            copy_utils[state_index] = [state_mapping[t[3]]] # cost for index (noted as copy) points to index of target
            iteration_utils[state_index] = np.nan # initialize as nan
        else:
            raise ValueError('Error: Bad utility type specification', u_type)
    
    # Fill copied utilities
    for u in copy_utils:
        iteration_utils[u] = iteration_utils[copy_utils[u]]
    if np.isnan(iteration_utils).any(): #TODO: Could make this output something more detailed. I.e. index of error utility
        raise ValueError('Error: NaN utility specified. Likely a copy type error. Please check utility sheet')

    print('calculating costs and utilities...')
    for state in state_mapping:
        idx = state_mapping[state]
        
        # TODO: add "division" of treatment states for cost or utility? How to effectivley do this...
        # The first if statement should never be triggered at the moment - just roughed in for future (assuming
        # understanding of application is correct)
        if state in multiple_toxicity_states:
            multiple_toxicity_states['state']
            results_log[:,idx,:] # is proportion of pts in that patient at every time iterations
            
            results_log_costs[:,idx,:] = results_log[:,idx,:] * iteration_costs[:,idx][:,np.newaxis] #costs_df.loc[costs_df.state==state].cost.values[0] #cost_array[row,iteration]
            results_log_utilities[:,idx,:] = results_log[:,idx,:] * iteration_utils[:,idx][:,np.newaxis] #utilities_df.loc[utilities_df.state==state].utility.values[0] #utility_array[row,iteration]
        else:
            # multiply every (proportion) entry with the cost/utility associated with that state at that iteration
            # assign result to utility/cost array
            results_log_costs[:,idx,:] = results_log[:,idx,:] * iteration_costs[:,idx][:,np.newaxis] #costs_df.loc[costs_df.state==state].cost.values[0] #cost_array[row,iteration]
            results_log_utilities[:,idx,:] = results_log[:,idx,:] * iteration_utils[:,idx][:,np.newaxis] #utilities_df.loc[utilities_df.state==state].utility.values[0] #utility_array[row,iteration]
    
    # Adjust utilities into per-year
    results_log_utilities = results_log_utilities * (cycle_length/365)
    
    # Apply discount rate
    # cost * (1 / ((1+discount_rate)**year))
    for i in range(0,results_log.shape[2]):
        year = math.floor((i*cycle_length)/365)
        results_log_costs[:,:,i] = results_log_costs[:,:,i] * (1 / ((1+discount_rate)**year))
        results_log_utilities[:,:,i] = results_log_utilities[:,:,i] * (1 / ((1+discount_rate)**year))

    print('done costs and utilities...')

    #---------------------------------------------------------------------------------------------------
    # Save/return result components in a raw form (ie. as 3D arrays)
    # 1) population
    # 2) costs
    # 3) utility
    #---------------------------------------------------------------------------------------------------
    if save:
        name = model_name + '_' + strftime("%d-%m-%Y")
        np.save('../model_outputs/'+name+'_costs.npy', results_log_costs)
        np.save('../model_outputs/'+name+'_utilities.npy', results_log_utilities)
        print('Cost and utility results saved: ../model_outputs/'+name+'... .npy')
    print('')

    return results_log, results_log_costs, results_log_utilities